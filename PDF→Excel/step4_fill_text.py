import openpyxl
from openpyxl.styles import Alignment
import re
import copy

def get_excel_geometry(ws):
    """Excelシートの全セル座標(相対)を取得"""
    row_heights = {}
    col_widths = {}
    max_row = ws.max_row
    max_col = ws.max_column

    total_height = 0
    row_starts = {1: 0}
    for r in range(1, max_row + 1):
        h = ws.row_dimensions[r].height
        if h is None: h = 13.5
        row_heights[r] = h
        total_height += h
        row_starts[r + 1] = total_height

    total_width = 0
    col_starts = {1: 0}
    for c in range(1, max_col + 1):
        col_letter = openpyxl.utils.get_column_letter(c)
        w = ws.column_dimensions[col_letter].width
        if w is None: w = 8.43
        w_points = w * 7.5
        col_widths[c] = w_points
        total_width += w_points
        col_starts[c + 1] = total_width

    # 結合セル情報のマップ
    merge_map = {}
    for rng in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = rng.min_col, rng.min_row, rng.max_col, rng.max_row
        x = col_starts[min_col]
        y = row_starts[min_row]
        w = col_starts[max_col + 1] - x
        h = row_starts[max_row + 1] - y
        rect = {
            "r_idx": min_row, "c_idx": min_col,
            "rel_x": x / total_width, "rel_y": y / total_height,
            "rel_w": w / total_width, "rel_h": h / total_height,
            "is_merged": True
        }
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                merge_map[(r, c)] = rect

    cells_geometry = []
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            if (r, c) in merge_map:
                rect = merge_map[(r, c)]
                if rect["r_idx"] == r and rect["c_idx"] == c:
                    cells_geometry.append(rect)
            else:
                x = col_starts[c]
                y = row_starts[r]
                w = col_widths[c]
                h = row_heights[r]
                cells_geometry.append({
                    "r_idx": r, "c_idx": c,
                    "rel_x": x / total_width, "rel_y": y / total_height,
                    "rel_w": w / total_width, "rel_h": h / total_height,
                    "is_merged": False
                })
    return cells_geometry

def split_box_elements(elements):
    """
    '□項目名' のようにボックスと文字が一体化している要素を
    ['□', '項目名'] の2つの要素に分割する
    """
    new_elements = []
    # ボックス文字（□、☐、■、☑）で始まり、その後に文字が続くパターン
    pattern = re.compile(r"^([□☐■☑])\s*(.+)$")

    for el in elements:
        if el["type"] != "text":
            new_elements.append(el)
            continue

        text = str(el["text"]).strip()
        match = pattern.match(text)
        
        if match:
            # 分割が必要
            box_char = match.group(1)
            remain_text = match.group(2)
            
            # 座標計算（ボックスは正方形に近いと仮定）
            height = el["bottom"] - el["top"]
            width_box = height * 1.2 # 少し余裕を持たせる
            
            # 1. ボックス要素
            el_box = copy.deepcopy(el)
            el_box["text"] = box_char
            el_box["x1"] = el["x0"] + width_box # 右端を縮める
            
            # 2. 残りのテキスト要素
            el_text = copy.deepcopy(el)
            el_text["text"] = remain_text
            el_text["x0"] = el["x0"] + width_box # 左端をずらす
            
            new_elements.append(el_box)
            new_elements.append(el_text)
            print(f"  [Split] '{text}' -> '{box_char}' & '{remain_text}'")
        else:
            new_elements.append(el)
            
    return new_elements

def fill_text_into_skeleton(skeleton_path, all_pages_elements, output_path):
    print(f"Loading skeleton: {skeleton_path} ...")
    wb = openpyxl.load_workbook(skeleton_path)

    for page_num, raw_elements in all_pages_elements.items():
        sheet_name = f"Page {page_num}"
        if sheet_name not in wb.sheetnames:
            print(f"[Skip] Sheet '{sheet_name}' not found.")
            continue
        ws = wb[sheet_name]
        print(f"Processing {sheet_name}...")

        # 1. Excelセル座標取得
        excel_cells = get_excel_geometry(ws)

        # 2. ページサイズ取得
        page_w, page_h = 1000, 1000
        for el in raw_elements:
            if el["type"] == "page_size":
                page_w, page_h = el["width"], el["height"]
                break
        
        # 3. テキスト分割処理（□の分離）
        elements = split_box_elements(raw_elements)

        # 4. 流し込み
        for el in elements:
            if el["type"] != "text": continue

            # 中心座標 (相対 0.0-1.0)
            cx_pdf = ((el["x0"] + el["x1"]) / 2) / page_w
            cy_pdf = ((el["top"] + el["bottom"]) / 2) / page_h

            best_cell = None
            min_dist = float("inf")

            for cell in excel_cells:
                cx_cell = cell["rel_x"] + cell["rel_w"] / 2
                cy_cell = cell["rel_y"] + cell["rel_h"] / 2
                
                # Y方向のズレをより厳しく判定 (*10)
                dist = ((cx_pdf - cx_cell) ** 2) + ((cy_pdf - cy_cell) ** 2) * 10
                if dist < min_dist:
                    min_dist = dist
                    best_cell = cell

            if best_cell:
                r, c = best_cell["r_idx"], best_cell["c_idx"]
                cell_obj = ws.cell(row=r, column=c)
                
                text_to_write = str(el["text"]).strip()
                if not text_to_write: continue

                # 既存テキストがある場合は改行で追記
                if cell_obj.value:
                    cell_obj.value = str(cell_obj.value) + "\n" + text_to_write
                else:
                    cell_obj.value = text_to_write
                
                cell_obj.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    wb.save(output_path)
    print(f"\nFinal Excel saved to: {output_path}")