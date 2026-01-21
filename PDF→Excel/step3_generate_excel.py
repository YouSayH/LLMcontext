import bisect
import re
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils.cell import get_column_letter

def clean_text(text):
    if text is None: return ""
    text = str(text).strip()
    text = re.sub(r'[\x00-\x09\x0b\x0c\x0e-\x1f\x7f]', '', text)
    if text.startswith("="):
        text = "'" + text
    return text

def cluster_coords(coords, tolerance):
    if not coords: return []
    sorted_coords = sorted(list(coords))
    clusters = []
    current_cluster = [sorted_coords[0]]
    for c in sorted_coords[1:]:
        if c - current_cluster[-1] <= tolerance:
            current_cluster.append(c)
        else:
            clusters.append(int(sum(current_cluster) / len(current_cluster)))
            current_cluster = [c]
    clusters.append(int(sum(current_cluster) / len(current_cluster)))
    return clusters

def apply_border_to_range(ws, r_start, c_start, r_end, c_end, border_type):
    """範囲全体の外枠に線を引く"""
    thin = Side(border_style="thin", color="000000")
    
    # 上辺
    if border_type in ['line_h', 'all', 'box']:
        for c in range(c_start, c_end + 1):
            cell = ws.cell(row=r_start, column=c)
            b = cell.border
            cell.border = Border(top=thin, left=b.left, right=b.right, bottom=b.bottom)
    
    # 下辺
    if border_type in ['line_h', 'all', 'box']:
        for c in range(c_start, c_end + 1):
            cell = ws.cell(row=r_end, column=c)
            b = cell.border
            cell.border = Border(top=b.top, left=b.left, right=b.right, bottom=thin)

    # 左辺
    if border_type in ['line_v', 'all', 'box']:
        for r in range(r_start, r_end + 1):
            cell = ws.cell(row=r, column=c_start)
            b = cell.border
            cell.border = Border(top=b.top, left=thin, right=b.right, bottom=b.bottom)

    # 右辺
    if border_type in ['line_v', 'all', 'box']:
        for r in range(r_start, r_end + 1):
            cell = ws.cell(row=r, column=c_end)
            b = cell.border
            cell.border = Border(top=b.top, left=b.left, right=thin, bottom=b.bottom)


def generate_excel(all_pages_elements, output_path, tolerance=30.0, dpi=300):
    wb = Workbook()
    if "Sheet" in wb.sheetnames: del wb["Sheet"]
    scale_factor = 72 / dpi 

    for page_num, elements in all_pages_elements.items():
        ws = wb.create_sheet(title=f"Page {page_num}")
        print(f"Generating Sheet: Page {page_num}...")

        # --- 1. グリッド座標の抽出 ---
        x_coords = set()
        y_coords = set()
        
        # ページサイズ
        page_w, page_h = 0, 0
        for el in elements:
            if el['type'] == 'page_size':
                page_w, page_h = el['width'], el['height']
                x_coords.add(0); x_coords.add(page_w)
                y_coords.add(0); y_coords.add(page_h)
        
        # 罫線座標
        lines = [el for el in elements if el['type'] in ['line_h', 'line_v']]
        for el in lines:
            if el['type'] == 'line_h':
                y_coords.add(el['top']); y_coords.add(el['bottom'])
                # 横線の端点も重要
                x_coords.add(el['x0']); x_coords.add(el['x1'])
            elif el['type'] == 'line_v':
                x_coords.add(el['x0']); x_coords.add(el['x1'])
                # 縦線の端点も重要
                y_coords.add(el['top']); y_coords.add(el['bottom'])

        cols = cluster_coords(x_coords, tolerance)
        rows = cluster_coords(y_coords, tolerance)
        
        print(f"  - Grid: {len(rows)} rows x {len(cols)} columns")

        # --- 2. 罫線の描画 ---
        # 罫線の位置にあるセルの境界に線を引く
        for el in lines:
            # 線がどのグリッド範囲をカバーしているか
            c_start = bisect.bisect_left(cols, el['x0'] + 2) - 1
            c_end   = bisect.bisect_left(cols, el['x1'] - 2) - 1
            r_start = bisect.bisect_left(rows, el['top'] + 2) - 1
            r_end   = bisect.bisect_left(rows, el['bottom'] - 2) - 1
            
            c_start = max(0, c_start)
            c_end = min(len(cols)-2, c_end)
            r_start = max(0, r_start)
            r_end = min(len(rows)-2, r_end)

            if c_end < c_start: c_end = c_start
            if r_end < r_start: r_end = r_start

            # Excelの行・列番号（1-based）
            # line_hなら上下、line_vなら左右に線を引く
            apply_border_to_range(ws, r_start+1, c_start+1, r_end+1, c_end+1, el['type'])

        # --- 3. テキストの配置 ---
        # 結合状態を管理（セットには (row, col) を入れる）
        occupied_cells = set()
        
        text_elements = sorted([e for e in elements if e['type'] == 'text'], key=lambda x: (x['top'], x['x0']))

        for el in text_elements:
            # テキストの中心点
            cx = (el['x0'] + el['x1']) / 2
            cy = (el['top'] + el['bottom']) / 2
            
            # 中心点が属するセルを特定
            r_idx = bisect.bisect_left(rows, cy) - 1
            c_idx = bisect.bisect_left(cols, cx) - 1
            
            if r_idx < 0 or r_idx >= len(rows)-1: continue
            if c_idx < 0 or c_idx >= len(cols)-1: continue
            
            excel_r = r_idx + 1
            excel_c = c_idx + 1

            # --- 結合範囲の自動拡張 ---
            # テキストの幅が現在のセル幅を超えていて、かつ「罫線をまたがない」なら結合する
            # つまり、隣のセルとの境界に物理的な罫線(lines)が存在しないかチェックする
            
            merge_c_end = c_idx
            
            # 右方向に拡張チェック
            current_right_x = cols[c_idx+1]
            target_right_x = el['x1']
            
            while current_right_x < target_right_x - tolerance:
                # 次の境界線 (cols[merge_c_end+1]) に縦線があるかチェック
                boundary_x = cols[merge_c_end+1]
                
                # このX座標付近に、現在の行の高さをカバーする縦線があるか？
                has_border = False
                for line in lines:
                    if line['type'] == 'line_v':
                        if abs(line['x0'] - boundary_x) < tolerance:
                            # 線のY範囲が、現在の行のY範囲と重なるか
                            l_top, l_btm = line['top'], line['bottom']
                            cell_top, cell_btm = rows[r_idx], rows[r_idx+1]
                            if not (l_btm < cell_top or l_top > cell_btm):
                                has_border = True
                                break
                
                if has_border:
                    break # 罫線があるのでこれ以上右には結合しない
                
                merge_c_end += 1
                if merge_c_end >= len(cols)-1: break
                current_right_x = cols[merge_c_end+1]
            
            excel_c_end = merge_c_end + 1
            
            # --- 結合実行 ---
            # 既存の結合と重ならないかチェック
            is_blocked = False
            for c in range(excel_c, excel_c_end + 1):
                if (excel_r, c) in occupied_cells:
                    is_blocked = True
                    break
            
            if (excel_c_end > excel_c) and (not is_blocked):
                try:
                    # openpyxlの既存結合チェック
                    conflict = False
                    for rng in ws.merged_cells.ranges:
                        if not (excel_r > rng.max_row or excel_r < rng.min_row or 
                                excel_c > rng.max_col or excel_c_end < rng.min_col):
                            conflict = True
                            break
                    if not conflict:
                        ws.merge_cells(start_row=excel_r, start_column=excel_c,
                                       end_row=excel_r, end_column=excel_c_end)
                        # 占有フラグ更新
                        for c in range(excel_c, excel_c_end + 1):
                            occupied_cells.add((excel_r, c))
                except: pass

            # --- 書き込み ---
            # 結合セルの場合、左上がターゲット。ただし既に占有されている場合は親を探す必要があるが
            # 上記ロジックでブロックしているので基本は左上でOK。
            # ただし、同じセルに複数のテキストが入る場合（改行など）の考慮が必要。
            
            # 書き込み先セル（結合されていれば親セル）
            target_cell = ws.cell(row=excel_r, column=excel_c)
            # もしこのセルが結合範囲の一部（左上以外）なら親を探す
            for rng in ws.merged_cells.ranges:
                if (rng.min_row <= excel_r <= rng.max_row) and (rng.min_col <= excel_c <= rng.max_col):
                    target_cell = ws.cell(row=rng.min_row, column=rng.min_col)
                    break
            
            txt = clean_text(el['text'])
            if target_cell.value:
                target_cell.value = str(target_cell.value) + "\n" + txt
            else:
                target_cell.value = txt
            
            target_cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')

        # --- 4. 列幅・行高 ---
        for i in range(len(cols) - 1):
            w_px = cols[i+1] - cols[i]
            # 係数を調整してExcelの見た目に近づける
            width = max(w_px * scale_factor * 0.11, 2)
            ws.column_dimensions[get_column_letter(i+1)].width = min(width, 80)

        for i in range(len(rows) - 1):
            h_px = rows[i+1] - rows[i]
            ws.row_dimensions[i+1].height = max(h_px * scale_factor * 0.75, 13.5)

    try:
        wb.save(output_path)
        print(f"\nSaved Excel to: {output_path}")
    except PermissionError:
        print(f"\n[Error] Could not save to {output_path}. Please close the file.")