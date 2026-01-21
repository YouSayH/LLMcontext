import bisect
import re
from openpyxl import Workbook
from openpyxl.styles import Border, Side
from openpyxl.utils.cell import get_column_letter

def cluster_coords(coords, tolerance=15):
    """座標をクラスタリングして整理"""
    if not coords:
        return []
    sorted_coords = sorted(list(coords))
    clusters = []
    current_cluster = [sorted_coords[0]]
    
    for c in sorted_coords[1:]:
        avg_val = sum(current_cluster) / len(current_cluster)
        if abs(c - avg_val) <= tolerance:
            current_cluster.append(c)
        else:
            clusters.append(int(sum(current_cluster) / len(current_cluster)))
            current_cluster = [c]
    
    clusters.append(int(sum(current_cluster) / len(current_cluster)))
    return clusters

def apply_border_safe(ws, r, c, border_type):
    """セルに罫線を引く"""
    thin = Side(border_style="thin", color="000000")
    try:
        cell = ws.cell(row=r, column=c)
        current = cell.border
        new_sides = {
            "left": current.left, "right": current.right,
            "top": current.top, "bottom": current.bottom
        }
        if border_type in ["line_h", "all", "box"]:
            new_sides["top"] = thin
            new_sides["bottom"] = thin
        if border_type in ["line_v", "all", "box"]:
            new_sides["left"] = thin
            new_sides["right"] = thin
        cell.border = Border(**new_sides)
    except:
        pass

def generate_skeleton(all_pages_elements, output_path, tolerance=50.0, dpi=300):
    wb = Workbook()
    if "Sheet" in wb.sheetnames: del wb["Sheet"]
    scale_factor = 72 / dpi
    
    # ボックス文字とみなす正規表現
    box_pattern = re.compile(r"^[□☐■☑]")

    for page_num, elements in all_pages_elements.items():
        ws = wb.create_sheet(title=f"Page {page_num}")
        print(f"Generating Skeleton with Boxes: Page {page_num}...")

        # ページサイズ
        page_w, page_h = 0, 0
        for el in elements:
            if el["type"] == "page_size":
                page_w, page_h = el["width"], el["height"]
                break
        if page_w == 0: page_w, page_h = 2480, 3508

        # --- 1. グリッド座標の抽出 ---
        x_coords = {0, page_w}
        y_coords = {0, page_h}
        
        # 線フィルタ（短すぎる線は無視）
        min_grid_line_w = page_w * 0.05
        min_grid_line_h = page_h * 0.05

        for el in elements:
            if el["type"] == "line_h":
                if (el["x1"] - el["x0"]) > min_grid_line_w:
                    y_coords.add(el["top"]); y_coords.add(el["bottom"])
                    x_coords.add(el["x0"]); x_coords.add(el["x1"])
            elif el["type"] == "line_v":
                if (el["bottom"] - el["top"]) > min_grid_line_h:
                    x_coords.add(el["x0"]); x_coords.add(el["x1"])
                    y_coords.add(el["top"]); y_coords.add(el["bottom"])
            
            # 【追加】テキスト内のボックス文字を検出して、グリッド候補に追加
            elif el["type"] == "text":
                text = str(el["text"]).strip()
                # "□"で始まる、かつ長さが短い（単体のボックス、またはボックス+短い文字）
                # OCRが「□ 項目」を1つに認識していても、左端(x0)と、ボックスの幅分右(x0 + height)を区切りとする
                if box_pattern.match(text):
                    # ボックスの左端は区切り候補
                    x_coords.add(el["x0"])
                    
                    # ボックスの右端（文字の高さとおおよそ同じ幅と仮定）
                    # これにより、ボックス直後に区切り線を入れる
                    box_width = (el["bottom"] - el["top"]) * 1.2 # 正方形より少し余裕を持つ
                    x_coords.add(el["x0"] + box_width)

        # クラスタリング
        cols = cluster_coords(x_coords, tolerance=tolerance)
        rows = cluster_coords(y_coords, tolerance=tolerance)

        print(f"  - Grid: {len(rows)-1} rows x {len(cols)-1} cols")

        # --- 2. 列幅・行高 ---
        for i in range(len(cols) - 1):
            w_px = cols[i+1] - cols[i]
            ws.column_dimensions[get_column_letter(i+1)].width = w_px * scale_factor * 0.12
        
        for i in range(len(rows) - 1):
            h_px = rows[i+1] - rows[i]
            ws.row_dimensions[i+1].height = h_px * scale_factor * 0.75

        # --- 3. 罫線描画 ---
        for el in elements:
            if el["type"] not in ["line_h", "line_v"]: continue
            
            # 座標合わせ
            c_start = bisect.bisect_left(cols, el["x0"] + tolerance/2) - 1
            c_end   = bisect.bisect_left(cols, el["x1"] - tolerance/2) - 1
            r_start = bisect.bisect_left(rows, el["top"] + tolerance/2) - 1
            r_end   = bisect.bisect_left(rows, el["bottom"] - tolerance/2) - 1

            c_start = max(0, min(c_start, len(cols)-2))
            c_end   = max(0, min(c_end, len(cols)-2))
            r_start = max(0, min(r_start, len(rows)-2))
            r_end   = max(0, min(r_end, len(rows)-2))

            if c_end < c_start: c_end = c_start
            if r_end < r_start: r_end = r_start

            for r in range(r_start, r_end + 1):
                for c in range(c_start, c_end + 1):
                    apply_border_safe(ws, r+1, c+1, el["type"])

    try:
        wb.save(output_path)
        print(f"\nSkeleton saved to: {output_path}")
    except PermissionError:
        print(f"\n[Error] Could not save to {output_path}.")