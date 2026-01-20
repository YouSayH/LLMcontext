import openpyxl
from openpyxl.cell.cell import MergedCell
import os
import argparse
import re
import glob

class ExcelMapper:
    """
    LLMがExcelシートの構造を理解するためのテキストマップを生成するクラス
    （連続する同一構造の行を圧縮する機能付き）
    """
    
    def __init__(self, file_path):
        self.file_path = file_path

    def _get_cell_features(self, cell):
        """
        セルの特徴（値、型、スタイル、入力欄らしさ）を分析してタグとテキストを返す
        """
        if isinstance(cell, MergedCell):
            return None 

        value = cell.value
        val_str = str(value).strip() if value is not None else ""
        
        tags = []

        # 枠線の確認
        has_border = False
        if cell.border:
            has_border = any([
                cell.border.top.style,
                cell.border.bottom.style,
                cell.border.left.style,
                cell.border.right.style
            ])
        
        # 背景色の確認
        has_color = False
        if cell.fill and cell.fill.start_color:
            try:
                color_index = cell.fill.start_color.index
                if color_index not in [None, '00000000', 64, 'FFFFFFFF']:
                    has_color = True
                    tags.append(f"<BG_COLOR:{str(color_index)[:2]}..>") 
            except:
                pass

        # --- LLMへのヒント生成 ---
        if val_str:
            if cell.is_date:
                tags.append("<DATE>")
            elif cell.data_type == 'f':
                tags.append("<FORMULA>")
            return f"{val_str} {' '.join(tags)}"
        else:
            if has_border:
                return "{{__INPUT_HERE__}} <BORDER>"
            elif has_color:
                return "{{__INPUT_HERE__}} <COLORED>"
            else:
                return "<EMPTY>"

    def _get_row_structure_signature(self, row_text):
        """
        行の構造比較用シグネチャを作成する。
        座標部分 [A1] などを削除し、構造が同じかどうか判定するための文字列を返す。
        """
        return re.sub(r'\[[A-Z]+[0-9]+\] ', '', row_text)

    def analyze_sheet(self, sheet_name):
        try:
            wb = openpyxl.load_workbook(self.file_path, data_only=True)
        except Exception as e:
            return f"Error loading {self.file_path}: {e}"

        if sheet_name not in wb.sheetnames:
            return f"Error: Sheet '{sheet_name}' not found."

        ws = wb[sheet_name]
        merged_ranges = [str(r) for r in ws.merged_cells.ranges]
        
        output_lines = []
        output_lines.append(f"=== SHEET MAP: {sheet_name} ===")
        output_lines.append(f"Merged Ranges: {', '.join(merged_ranges)}")
        output_lines.append("Legend: {__INPUT_HERE__} = Likely an input field (Empty + Border/Color)")
        output_lines.append("-" * 60)

        max_r = min(ws.max_row, 300) 
        max_c = min(ws.max_column, 50)

        # 生の行データを生成
        raw_rows = []
        for row in ws.iter_rows(min_row=1, max_row=max_r, min_col=1, max_col=max_c):
            row_cells_text = []
            has_meaningful_content = False
            
            for cell in row:
                coord = cell.coordinate
                cell_text = self._get_cell_features(cell)
                
                if cell_text is None:
                    continue 
                
                if cell_text == "<EMPTY>":
                    pass
                else:
                    row_cells_text.append(f"[{coord}] {cell_text}")
                    has_meaningful_content = True

            if has_meaningful_content:
                raw_rows.append(" | ".join(row_cells_text))
            else:
                raw_rows.append("<EMPTY_ROW>")

        # --- 圧縮ロジック (同一構造行の省略) ---
        if not raw_rows:
            return "Empty Sheet"

        compressed_lines = []
        
        current_line = raw_rows[0]
        current_sig = self._get_row_structure_signature(current_line)
        repeat_count = 0

        for i in range(1, len(raw_rows)):
            next_line = raw_rows[i]
            next_sig = self._get_row_structure_signature(next_line)

            if next_line == "<EMPTY_ROW>" and current_line == "<EMPTY_ROW>":
                 repeat_count += 1
            elif next_line != "<EMPTY_ROW>" and next_sig == current_sig:
                repeat_count += 1
            else:
                compressed_lines.append(current_line)
                if repeat_count > 0:
                    compressed_lines.append(f"... (Identical structure repeated for {repeat_count} rows) ...")
                
                current_line = next_line
                current_sig = next_sig
                repeat_count = 0

        compressed_lines.append(current_line)
        if repeat_count > 0:
            compressed_lines.append(f"... (Identical structure repeated for {repeat_count} rows to end) ...")

        output_lines.extend(compressed_lines)
        wb.close()
        
        return "\n".join(output_lines)

def analyze_excel_for_llm(file_path, sheet_name=None):
    mapper = ExcelMapper(file_path)
    if sheet_name:
        return mapper.analyze_sheet(sheet_name)
    else:
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            results = []
            for name in wb.sheetnames:
                results.append(mapper.analyze_sheet(name))
            wb.close()
            return "\n\n".join(results)
        except Exception as e:
            return f"Error: {e}"

# --- Main Execution ---
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate Excel Map for LLM")
    # 引数を任意(nargs='?')にし、デフォルト値を削除
    parser.add_argument("filename", nargs="?", help="Target Excel filename")
    parser.add_argument("sheetname", nargs="?", help="Target Sheet name")
    args = parser.parse_args()

    if args.filename:
        # ファイル指定がある場合
        if os.path.exists(args.filename):
            print(f"--- Analyzing File: {args.filename} ---")
            print(analyze_excel_for_llm(args.filename, args.sheetname))
        else:
            print(f"File {args.filename} not found.")
    else:
        # 引数がない場合 -> カレントディレクトリの全xlsxファイルをスキャン
        xlsx_files = glob.glob("*.xlsx")
        
        if not xlsx_files:
            print("No .xlsx files found in the current directory.")
        else:
            print(f"Found {len(xlsx_files)} Excel files. Scanning all...\n")
            for f in xlsx_files:
                print("============================================================")
                print(f" FILE: {f}")
                print("============================================================")
                print(analyze_excel_for_llm(f))
                print("\n")